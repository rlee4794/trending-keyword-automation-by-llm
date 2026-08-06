#!/usr/bin/env python3
"""Export daily_trending_{REGION}.json to a formatted XLSX workbook.

Produces a 3-sheet workbook:
  1. 本日要點 — daily highlights with signal + description
  2. 熱門菜式 — Top N dish keywords with term→concept, posts, likes, background
  3. 原始數據摘要 — pipeline execution summary

Background column: 50-70 char CJK summary, NOT raw caption excerpt.
Format: [{source}] {venue} {context}，{feature/reaction}
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# ── helpers ──────────────────────────────────────────────────────────

TITLE_FONT = Font(name="Microsoft YaHei", size=16, bold=True, color="1F4E79")
HEADER_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
BODY_FONT = Font(name="Microsoft YaHei", size=10)
SUBTITLE_FONT = Font(name="Microsoft YaHei", size=9, color="666666")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

REGION_LABEL = {"hk": "HK", "tw": "TW"}

# Context detection
CONTEXT_MAP = {
    "聯乘": "聯乘", "×": "聯乘", "聯名": "聯乘", "EVA": "聯乘", "Chiikawa": "聯乘",
    "新開": "新開", "全新": "新開", "開幕": "新開", "新店": "新開", "進駐": "新開",
    "限定": "限定", "期間限定": "限定",
    "任食": "任食", "放題": "放題", "buffet": "任食",
    "米芝蓮": "米芝蓮", "Michelin": "米芝蓮",
    "老字號": "老字號", "40年": "老字號", "50年": "老字號",
    "酒店": "酒店級", "五星": "酒店級",
    "創意": "創意", "首創": "創意",
    "限時": "限時", "優惠": "優惠",
    "抵食": "抵食", "CP值": "高CP", "性價比": "高CP",
    "平民": "平民", "街坊": "街坊",
    "人氣": "人氣", "排隊": "排隊",
    "隱世": "隱世",
    "新品": "新品", "新出": "新品",
}

# Phrases that indicate non-food social media CTA — skip these
CTA_SKIP = [
    "快啲Share比朋友", "快啲Follow", "快啲留言", "留言抽", "留言「", 
    "記得Follow", "記得Like", "Share俾朋友", "Tag你朋友", "tag你",
    "傳給你想", "傳給你的", "Save低", "Bookmark", "即刻send俾",
    "想睇更多", "密切留意", "記得去", "快啲去試",
    "Follow us", "credit：", "credit:", "DM我哋",
    "你願意唔願意", "希望幫到你", "加油",
]


def _style_header_row(ws, row: int, headers: list[str]) -> None:
    for col, val in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _write_cell(ws, row: int, col: int, value, *, bold: bool = False, center: bool = False) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Microsoft YaHei", size=10, bold=bold)
    cell.border = THIN_BORDER
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center",
        wrap_text=True,
    )


# ── background builder ───────────────────────────────────────────────

def _extract_venue(caption: str) -> str:
    """Extract restaurant/venue name from caption."""
    # 【店名】
    m = re.search(r'【(.+?)】', caption)
    if m:
        name = m.group(1).strip()
        if 2 <= len(name) <= 25 and not name.startswith("#"):
            return name
    # 📍 @handle
    m = re.search(r'📍\s*@(\S+)', caption)
    if m:
        return m.group(1).strip()
    # Known brand detection from caption
    brands = ["肯德基", "KFC", "IKEA", "宜家", "麥當勞", "McDonald", "Bakehouse",
              "牛大人", "喜茶", "HEYTEA", "壽司郎", "PICI", "牛惣居酎屋"]
    for brand in brands:
        if brand in caption:
            return brand
    return ""


def _infer_context(caption: str) -> str:
    """Infer context labels from caption content."""
    found = []
    for keyword, label in CONTEXT_MAP.items():
        if keyword in caption and label not in found:
            found.append(label)
    return "、".join(found)


def _is_cta(phrase: str) -> bool:
    """Check if a phrase is a social media call-to-action."""
    for skip in CTA_SKIP:
        if skip in phrase:
            return True
    # Also skip very short or hashtag-heavy phrases
    if len(phrase) < 8:
        return True
    if phrase.count("#") >= 2:
        return True
    return False


def _extract_feature(caption: str, keyword_term: str) -> str:
    """Extract a short food-description phrase."""
    cap_clean = caption.replace("\n", " ").strip()
    phrases = [p.strip() for p in cap_clean.split("。") if p.strip()]
    
    # Skip phrases
    skip_words = ["地址", "營業時間", "電話", "IG：", "官方IG", "credit", "圖片來源",
                  "想睇更多", "記得Follow", "快啲Share", "留言抽", "傳給你想"]
    
    # Prefer phrases containing the keyword
    for phrase in phrases:
        if _is_cta(phrase):
            continue
        if any(s in phrase for s in skip_words):
            continue
        if keyword_term in phrase and 8 <= len(phrase) <= 55:
            return phrase
    
    # Fallback
    for phrase in phrases:
        if _is_cta(phrase):
            continue
        if any(s in phrase for s in skip_words):
            continue
        if 8 <= len(phrase) <= 55:
            return phrase
    
    return ""


def _pick_source_label(sources: list[str]) -> str:
    """Pick the most informative source: prefer user handles over hashtags."""
    users = [s for s in sources if s.startswith("@")]
    if users:
        return users[0]
    # Fall back to first non-generic hashtag
    generic = {"hkfood", "香港美食", "hkfoodie", "hkig", "hk", "香港", "hongkong"}
    good = [s for s in sources if s.lower().replace("#", "") not in generic and len(s) > 3]
    if good:
        return good[0]
    return ""


def _build_background(kw: dict, post_indices: list[int], posts: list[dict]) -> str:
    """Build a 50-70 char background summary."""
    term = kw.get("term", "")
    sources = kw.get("sources", [])
    source_label = _pick_source_label(sources)
    
    # Collect all captions
    captions = []
    for idx in post_indices:
        if idx < len(posts):
            cap = posts[idx].get("caption_snippet", "") or ""
            if cap:
                captions.append(cap)
    
    combined = " ".join(captions)
    
    # Try to extract venue + context from captions
    venue = _extract_venue(combined) if captions else ""
    context = _infer_context(combined) if captions else ""
    feature = _extract_feature(combined, term) if captions else ""
    
    # Build parts
    parts = []
    if source_label:
        parts.append(f"[{source_label}]")
    if venue:
        parts.append(venue)
    if context:
        parts.append(context)
    
    if feature:
        # Clean up feature: remove leading numbers/dashes, truncate
        feature = re.sub(r'^[\d\-\s•\.\、]+', '', feature).strip()
        if len(feature) > 50:
            feature = feature[:47] + "…"
        parts.append(feature)
    
    bg = "，".join(parts) if parts else f"{kw.get('post_count', 0)} 帖提及"
    
    # Trim to 70 CJK chars
    if len(bg) > 70:
        bg = bg[:67] + "…"
    
    return bg


# ── signal inference ─────────────────────────────────────────────────

def _infer_highlights(keywords: list[dict]) -> list[tuple[str, str]]:
    highlights: list[tuple[str, str]] = []
    used = set()

    for kw in keywords:
        term = kw.get("term", "")
        concept = kw.get("concept", term)
        likes = kw.get("total_likes", 0)
        post_count = kw.get("post_count", 0)
        sources = kw.get("sources", [])
        source_str = ", ".join(sources[:3]) if sources else ""

        display = f"{term} → {concept}" if term != concept else term

        if post_count == 1 and likes > 8000 and display not in used:
            highlights.append((f"{display} 單帖爆發",
                f"單帖 {likes/1000:.1f}K likes{('，來自 ' + source_str) if source_str else ''}，極高互動"))
            used.add(display)

        if post_count >= 4 and likes > 10000 and display not in used:
            highlights.append((f"{display} 持續熱度",
                f"{post_count} 帖累計 {likes/1000:.1f}K likes，跨平台/跨 source 持續發酵"))
            used.add(display)

    if len(highlights) < 3:
        total_posts = sum(k.get("post_count", 0) for k in keywords)
        total_keywords = len(keywords)
        if "pipeline_stats" not in used:
            highlights.append(("Pipeline 摘要",
                f"共 {total_posts} posts 通過 threshold，提取 {total_keywords} 個關鍵詞"))
            used.add("pipeline_stats")

    return highlights[:5]


# ── sheet builders ───────────────────────────────────────────────────

def _build_highlights_sheet(wb: Workbook, highlights: list[tuple[str, str]], date: str, region: str) -> None:
    ws = wb.active
    ws.title = "本日要點"
    region_label = REGION_LABEL.get(region, region.upper())

    ws.merge_cells("A1:B1")
    ws["A1"] = f"📊 {region_label} F&B 熱門關鍵字 — 本日要點（{date}）"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:B2")
    ws["A2"] = "Pipeline 自動生成 | 資料來源: Instagram + Threads | Google Trends: disabled"
    ws["A2"].font = SUBTITLE_FONT
    ws.row_dimensions[2].height = 22

    _style_header_row(ws, 4, ["要點", "說明"])
    ws.row_dimensions[4].height = 25

    for i, (label, desc) in enumerate(highlights, 5):
        _write_cell(ws, i, 1, label, bold=True)
        _write_cell(ws, i, 2, desc)
        ws.row_dimensions[i].height = 30

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 70


def _build_dishes_sheet(wb: Workbook, keywords: list[dict], posts: list[dict], date: str, region: str, top_n: int) -> None:
    ws = wb.create_sheet("熱門菜式")
    region_label = REGION_LABEL.get(region, region.upper())

    ws.merge_cells("A1:F1")
    ws["A1"] = f"🔥 Social 熱門菜式（按互動熱度，Top {top_n}）"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:F2")
    ws["A2"] = f"資料來源: Instagram + Threads | 日期: {date} | 地區: {region_label}"
    ws["A2"].font = SUBTITLE_FONT

    _style_header_row(ws, 4, ["關鍵詞", "Concept", "Posts", "Likes", "Shares", "背景"])

    for k in keywords:
        k["engagement"] = k.get("total_likes", 0) + k.get("total_shares", 0)
    active = sorted(
        [k for k in keywords if k.get("post_count", 0) > 0],
        key=lambda x: x["engagement"],
        reverse=True,
    )[:top_n]

    for i, kw in enumerate(active, 5):
        term = kw.get("term", "")
        concept = kw.get("concept", term)
        post_count = kw.get("post_count", 0)
        likes = kw.get("total_likes", 0)
        shares = kw.get("total_shares", 0)
        post_indices = kw.get("post_indices", [])

        display = f"{term} → {concept}" if term != concept else term

        bg = kw.get("background", "")
        if not bg:
            bg = _build_background(kw, post_indices, posts)

        _write_cell(ws, i, 1, display)
        _write_cell(ws, i, 2, concept)
        _write_cell(ws, i, 3, post_count, center=True)
        _write_cell(ws, i, 4, f"{likes/1000:.1f}K", center=True)
        _write_cell(ws, i, 5, f"{shares/1000:.1f}K", center=True)
        _write_cell(ws, i, 6, bg)
        ws.row_dimensions[i].height = 35

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 55


def _build_summary_sheet(wb: Workbook, data: dict, date: str, region: str) -> None:
    ws = wb.create_sheet("原始數據摘要")

    ws.merge_cells("A1:C1")
    ws["A1"] = "Pipeline 執行摘要"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 30

    threshold = data.get("threshold", {})
    ig_th = threshold.get("instagram", {})
    threads_th = threshold.get("threads", {})

    data_posts = data.get("posts", [])
    ig_count = sum(1 for p in data_posts if p.get("platform") == "instagram")
    threads_count = sum(1 for p in data_posts if p.get("platform") == "threads")
    active_kw = sum(1 for k in data.get("keywords", []) if k.get("post_count", 0) > 0)

    summary = [
        ("日期", date),
        ("地區", REGION_LABEL.get(region, region.upper())),
        ("通過 threshold posts", f"{len(data_posts)} ({ig_count} IG + {threads_count} Threads)"),
        ("提取 keyword 數", f"{active_kw} (有 post 關聯)"),
        ("Threshold (IG)", f"like≥{ig_th.get('min_likes', '?')} OR share≥{ig_th.get('min_shares', '?')}"),
        ("Threshold (Threads)", f"like≥{threads_th.get('min_likes', '?')} OR share≥{threads_th.get('min_shares', '?')}"),
        ("Google Trends", "enabled" if threshold.get("google", {}).get("enabled") else "disabled"),
        ("Generated at", data.get("generated_at", "")),
    ]

    for i, (k, v) in enumerate(summary, 3):
        _write_cell(ws, i, 1, k, bold=True)
        _write_cell(ws, i, 2, v)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 55


# ── main ─────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="Export daily trending data to XLSX")
    parser.add_argument("--date", required=True, help="Date (YYYY-MM-DD)")
    parser.add_argument("--region", default="hk", choices=["hk", "tw"], help="Region (default: hk)")
    parser.add_argument("--output", help="Output XLSX path (default: workspace)")
    parser.add_argument("--top", type=int, default=30, help="Top N dishes to include (default: 30)")
    args = parser.parse_args()

    skill_dir = Path(__file__).resolve().parent.parent
    json_path = skill_dir / "runs" / args.date / f"daily_trending_{args.region.upper()}.json"

    if not json_path.exists():
        print(f"ERROR: {json_path} not found. Run the pipeline first.", file=sys.stderr)
        sys.exit(1)

    with open(json_path) as f:
        data = json.load(f)

    keywords = data.get("keywords", [])
    posts = data.get("posts", [])

    highlights = _infer_highlights(keywords)

    wb = Workbook()
    _build_highlights_sheet(wb, highlights, args.date, args.region)
    _build_dishes_sheet(wb, keywords, posts, args.date, args.region, args.top)
    _build_summary_sheet(wb, data, args.date, args.region)

    if args.output:
        output = Path(args.output)
    else:
        output = Path.home() / ".openclaw" / "workspace" / f"{REGION_LABEL.get(args.region, args.region.upper())}_FB_Trending_{args.date}.xlsx"

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output))
    print(f"✅ XLSX exported: {output}")


if __name__ == "__main__":
    main()
